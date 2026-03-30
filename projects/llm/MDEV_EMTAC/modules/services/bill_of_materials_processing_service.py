from __future__ import annotations

import os
import shutil
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.orm.exc import NoResultFound

from modules.configuration.config import (
    DB_LOADSHEET_BOMS,
    DATABASE_PATH_IMAGES_FOLDER,
)
from modules.configuration.log_config import logger
from modules.emtacdb.emtacdb_fts import (
    Image,
    Part,
    PartsPositionImageAssociation,
)


class BillOfMaterialsProcessingService:
    """
    Service layer for BOM workbook processing and part/image/position associations.

    HARD RULES:
    - NEVER create sessions
    - NEVER commit
    - NEVER rollback
    - NEVER close sessions
    - Orchestrator owns transaction lifecycle

    LEGACY PARITY GOALS:
    - Preserve BOM workbook processing flow
    - Preserve target workbook naming behavior
    - Preserve BOM/photo matching behavior
    - Preserve fallback image search behavior
    - Preserve part-position-image association behavior
    """

    ALLOWED_BOM_EXTENSIONS = {"xlsx"}

    # ==========================================================
    # BASIC FILE VALIDATION
    # ==========================================================
    def allowed_file(self, filename: str) -> bool:
        """
        Validate uploaded BOM file name.

        Preserves legacy behavior:
        - only allows .xlsx
        """
        logger.info("Checking if file is allowed | filename=%s", filename)
        allowed = "." in filename and filename.rsplit(".", 1)[1].lower() in self.ALLOWED_BOM_EXTENSIONS
        logger.info("File allowed check complete | filename=%s | allowed=%s", filename, allowed)
        return allowed

    # ==========================================================
    # TARGET WORKBOOK PREP
    # ==========================================================
    def prompt_for_target_file(self, source_file: str) -> str:
        """
        Build the target workbook path used by the legacy BOM pipeline.

        Legacy naming preserved:
        bom_for_ABC.xlsx -> load_bom_for_ABC.xlsx
        """
        logger.info("Generating target workbook path | source_file=%s", source_file)

        suffix = os.path.basename(source_file).replace("bom_for_", "").replace(".xlsx", "")

        if not os.path.exists(DB_LOADSHEET_BOMS):
            os.makedirs(DB_LOADSHEET_BOMS, exist_ok=True)
            logger.info("Created BOM loadsheet directory | path=%s", DB_LOADSHEET_BOMS)

        target_file_name = f"load_bom_for_{suffix}.xlsx"
        file_path = os.path.join(DB_LOADSHEET_BOMS, target_file_name)

        if os.path.exists(file_path):
            logger.info("Target workbook already exists | target_path=%s", file_path)
        else:
            logger.info("Target workbook will be created | target_path=%s", file_path)

        return file_path

    def copy_bom_sheet_to_target(self, source_path: str, target_path: str) -> None:
        """
        Copy the BOM sheet from the uploaded workbook into the target workbook.

        Preserves legacy behavior:
        - source sheet must be named 'BOM'
        - new sheet name becomes bom_<suffix>
        - creates part_position_image sheet with fixed headers
        """
        logger.info("Copying BOM sheet | source_path=%s | target_path=%s", source_path, target_path)

        wb_source = load_workbook(source_path)

        if "BOM" not in wb_source.sheetnames:
            logger.error("Source workbook missing BOM sheet | source_path=%s", source_path)
            raise KeyError(f"The source workbook does not contain a sheet named 'BOM': {source_path}")

        bom_sheet = wb_source["BOM"]

        if os.path.exists(target_path):
            wb_target = load_workbook(target_path)
        else:
            wb_target = Workbook()
            default_sheet = wb_target.active
            wb_target.remove(default_sheet)

        suffix = os.path.basename(target_path).replace("load_bom_for_", "").replace(".xlsx", "")
        new_bom_sheet_name = f"bom_{suffix}"

        if new_bom_sheet_name in wb_target.sheetnames:
            logger.info(
                "Removing existing target BOM sheet before recreating | sheet=%s",
                new_bom_sheet_name,
            )
            del wb_target[new_bom_sheet_name]

        bom_target_sheet = wb_target.create_sheet(new_bom_sheet_name)

        for row in bom_sheet.iter_rows(values_only=True):
            bom_target_sheet.append(row)

        if "part_position_image" in wb_target.sheetnames:
            logger.info("Removing existing part_position_image sheet before recreating")
            del wb_target["part_position_image"]

        part_position_image_sheet = wb_target.create_sheet("part_position_image")
        part_position_image_sheet.append(["part", "position", "image", "description"])

        wb_target.save(target_path)
        logger.info(
            "Copied BOM sheet to target workbook | target_path=%s | bom_sheet=%s",
            target_path,
            new_bom_sheet_name,
        )

    # ==========================================================
    # MAIN PROCESSOR
    # ==========================================================
    def process_bom_loadsheet(
        self,
        session: Session,
        file_path: str,
        image_path: str,
        position_id: int,
    ) -> str:
        """
        Main BOM processing workflow.

        Preserves the original processing order:
        1. Build target workbook path
        2. Copy BOM sheet into target workbook
        3. Attempt part_list_image matching
        4. Search image folders for remaining items

        Returns:
            target_path: the generated target workbook path
        """
        logger.info(
            "Starting BOM loadsheet process | file_path=%s | image_path=%s | position_id=%s",
            file_path,
            image_path,
            position_id,
        )

        target_path = self.prompt_for_target_file(file_path)
        self.copy_bom_sheet_to_target(file_path, target_path)
        self.match_items_in_part_list_image(
            session=session,
            target_path=target_path,
            image_path=image_path,
            position_id=position_id,
        )
        self.process_remaining_items_in_image_folder(
            session=session,
            target_path=target_path,
            image_path=image_path,
            position_id=position_id,
        )

        logger.info(
            "Completed BOM loadsheet process | file_path=%s | target_path=%s | position_id=%s",
            file_path,
            target_path,
            position_id,
        )
        return target_path

    # ==========================================================
    # MATCH AGAINST part_list_image.xlsx
    # ==========================================================
    def match_items_in_part_list_image(
        self,
        session: Session,
        target_path: str,
        image_path: str,
        position_id: int,
    ) -> None:
        """
        Match BOM rows against part_list_image.xlsx / photo_list sheet.

        Preserves legacy behavior:
        - BOM item number from 4th column (row[3])
        - compare item_number[1:7] against first 6 chars of photo_list part number
        - process up to 3 photos
        """
        try:
            logger.info(
                "Matching BOM items against part_list_image workbook | target_path=%s | image_path=%s | position_id=%s",
                target_path,
                image_path,
                position_id,
            )

            wb_target = load_workbook(target_path)
            bom_sheet_name = self._get_bom_sheet_name(wb_target)
            bom_sheet = wb_target[bom_sheet_name]
            _ = wb_target["part_position_image"]  # preserved for legacy parity / validation

            part_list_image_path = os.path.join(DB_LOADSHEET_BOMS, "part_list_image.xlsx")
            wb_part_list = load_workbook(part_list_image_path)
            photo_list_sheet = wb_part_list["photo_list"]

            for row_idx, row in enumerate(bom_sheet.iter_rows(min_row=2, values_only=True), start=2):
                item_number = self._safe_str(row[3]).strip()
                logger.info(
                    "Processing BOM item for photo list match | row=%s | item_number=%s",
                    row_idx,
                    item_number,
                )

                if not item_number:
                    logger.warning("Skipping empty item number in BOM row | row=%s", row_idx)
                    continue

                part_number_prefix = item_number[1:7].upper()
                match_found = False

                for photo_row in photo_list_sheet.iter_rows(min_row=2, values_only=True):
                    photo_part_number_prefix = self._safe_str(photo_row[0])[:6].strip().upper()

                    logger.debug(
                        "Comparing BOM prefix to photo list prefix | bom_prefix=%s | photo_prefix=%s",
                        part_number_prefix,
                        photo_part_number_prefix,
                    )

                    if part_number_prefix == photo_part_number_prefix:
                        logger.info(
                            "Match found in part_list_image | row=%s | item_number=%s | prefix=%s",
                            row_idx,
                            item_number,
                            part_number_prefix,
                        )

                        photo_a = photo_row[1]
                        photo_b = photo_row[2]
                        photo_c = photo_row[3]

                        # Legacy index behavior preserved exactly
                        desc_a = photo_row[4] if len(photo_row) > 4 else None
                        desc_b = photo_row[5] if len(photo_row) > 5 else None
                        desc_c = photo_row[6] if len(photo_row) > 6 else None
                        manufacturer_description = photo_row[7] if len(photo_row) > 7 else None

                        prefixed_photo_a = f"{item_number[:1]}{photo_a}" if photo_a else None
                        prefixed_photo_b = f"{item_number[:1]}{photo_b}" if photo_b else None
                        prefixed_photo_c = f"{item_number[:1]}{photo_c}" if photo_c else None

                        if prefixed_photo_a:
                            logger.info(
                                "Processing matched photo A | item_number=%s | photo=%s | desc=%s | manufacturer_description=%s",
                                item_number,
                                prefixed_photo_a,
                                desc_a,
                                manufacturer_description,
                            )
                            self.process_part_position_image(
                                session=session,
                                part_number=item_number,
                                position_id=position_id,
                                image_title=prefixed_photo_a,
                                base_image_path=image_path,
                            )

                        if prefixed_photo_b:
                            logger.info(
                                "Processing matched photo B | item_number=%s | photo=%s | desc=%s | manufacturer_description=%s",
                                item_number,
                                prefixed_photo_b,
                                desc_b,
                                manufacturer_description,
                            )
                            self.process_part_position_image(
                                session=session,
                                part_number=item_number,
                                position_id=position_id,
                                image_title=prefixed_photo_b,
                                base_image_path=image_path,
                            )

                        if prefixed_photo_c:
                            logger.info(
                                "Processing matched photo C | item_number=%s | photo=%s | desc=%s | manufacturer_description=%s",
                                item_number,
                                prefixed_photo_c,
                                desc_c,
                                manufacturer_description,
                            )
                            self.process_part_position_image(
                                session=session,
                                part_number=item_number,
                                position_id=position_id,
                                image_title=prefixed_photo_c,
                                base_image_path=image_path,
                            )

                        match_found = True
                        break

                if not match_found:
                    logger.info(
                        "No match found in part_list_image for BOM row | row=%s | item_number=%s",
                        row_idx,
                        item_number,
                    )

            wb_target.save(target_path)
            logger.info("Saved target workbook after part_list_image matching | target_path=%s", target_path)

        except FileNotFoundError as exc:
            logger.error("part_list_image file not found | error=%s", exc)
        except KeyError as exc:
            logger.error("Workbook sheet lookup failed | error=%s", exc)
        except Exception as exc:
            logger.exception("Unexpected error during part_list_image matching | error=%s", exc)
            raise

    # ==========================================================
    # SUPPORTING ROW UTILITIES
    # ==========================================================
    def process_row(
        self,
        part_position_image_sheet,
        item_number: str,
        photo: Optional[str],
        description: Optional[str],
        manufacturer_description: Optional[str],
        position_id: int,
    ) -> None:
        """
        Preserved legacy helper for writing rows into part_position_image sheet.
        """
        logger.info(
            "Preparing workbook row append | item_number=%s | photo=%s | description=%s | manufacturer_description=%s | position_id=%s",
            item_number,
            photo,
            description,
            manufacturer_description,
            position_id,
        )

        if photo:
            full_description = (
                f"{description}, {manufacturer_description}"
                if manufacturer_description
                else description
            )
            part_position_image_sheet.append([item_number, position_id, photo, full_description])
            logger.info(
                "Workbook row appended | item_number=%s | position_id=%s | photo=%s | description=%s",
                item_number,
                position_id,
                photo,
                full_description,
            )
        else:
            logger.warning("No photo provided, row not appended | item_number=%s", item_number)

    def find_image_in_subfolders(self, image_title: str, base_path: str) -> Optional[str]:
        """
        Search for an image by filename-without-extension in base_path and all subfolders.
        """
        logger.info("Searching for image in subfolders | image_title=%s | base_path=%s", image_title, base_path)

        if not base_path or not os.path.exists(base_path):
            logger.warning("Base image path does not exist | base_path=%s", base_path)
            return None

        for root, _, files in os.walk(base_path):
            logger.debug("Searching directory | root=%s", root)
            for file in files:
                file_name_without_ext = os.path.splitext(file)[0]
                if file_name_without_ext == image_title:
                    full_path = os.path.join(root, file)
                    logger.info(
                        "Found matching image | image_title=%s | full_path=%s",
                        image_title,
                        full_path,
                    )
                    return full_path

        logger.warning(
            "Image not found in base path or subfolders | image_title=%s | base_path=%s",
            image_title,
            base_path,
        )
        return None

    # ==========================================================
    # FALLBACK SEARCH IN IMAGE FOLDER
    # ==========================================================
    def process_remaining_items_in_image_folder(
        self,
        session: Session,
        target_path: str,
        image_path: str,
        position_id: int,
    ) -> None:
        """
        Process BOM items not matched by part_list_image by searching image folders.

        Preserves legacy behavior:
        - uses first 7 characters of BOM item number
        - if no image found, creates part-position association without image
        """
        logger.info(
            "Processing remaining BOM items in image folder | target_path=%s | image_path=%s | position_id=%s",
            target_path,
            image_path,
            position_id,
        )

        wb_target = load_workbook(target_path)
        bom_sheet_name = self._get_bom_sheet_name(wb_target)
        bom_sheet = wb_target[bom_sheet_name]
        _ = wb_target["part_position_image"]  # preserved for legacy parity / validation

        for row_idx, row in enumerate(bom_sheet.iter_rows(min_row=2, values_only=True), start=2):
            item_number = self._safe_str(row[3])

            logger.info("Processing remaining BOM item | row=%s | item_number=%s", row_idx, item_number)

            if not item_number:
                logger.warning("Skipping empty item number during remaining-item search | row=%s", row_idx)
                continue

            first_seven_chars = item_number[:7]
            logger.info(
                "Attempting fallback image match | row=%s | item_number=%s | search_key=%s",
                row_idx,
                item_number,
                first_seven_chars,
            )

            image_file_path = self.find_image_in_subfolders(first_seven_chars, image_path)

            if image_file_path:
                logger.info(
                    "Fallback image match found | row=%s | item_number=%s | image_file_path=%s",
                    row_idx,
                    item_number,
                    image_file_path,
                )
                self.process_part_position_image(
                    session=session,
                    part_number=item_number,
                    position_id=position_id,
                    image_title=first_seven_chars,
                    base_image_path=image_path,
                )
            else:
                logger.warning(
                    "No fallback image found; creating part-position entry with no image | row=%s | item_number=%s",
                    row_idx,
                    item_number,
                )
                self.create_part_position_entry_no_image(
                    session=session,
                    part_number=item_number,
                    position_id=position_id,
                )

        wb_target.save(target_path)
        logger.info("Saved target workbook after fallback image search | target_path=%s", target_path)

    # ==========================================================
    # DB ASSOCIATION HELPERS
    # ==========================================================
    def create_part_position_entry_no_image(
        self,
        session: Session,
        part_number: str,
        position_id: int,
    ) -> Optional[PartsPositionImageAssociation]:
        """
        Create part_position_image association with no image.

        Returns the created association or None if the part is not found.
        """
        normalized_part_number = part_number[:7]
        logger.info(
            "Creating part-position association without image | part_number=%s | normalized_part_number=%s | position_id=%s",
            part_number,
            normalized_part_number,
            position_id,
        )

        try:
            part = session.query(Part).filter(Part.part_number == normalized_part_number).one()
        except NoResultFound:
            logger.error("No part found for no-image association | part_number=%s", part_number)
            return None

        existing_association = (
            session.query(PartsPositionImageAssociation)
            .filter(
                PartsPositionImageAssociation.part_id == part.id,
                PartsPositionImageAssociation.position_id == position_id,
                PartsPositionImageAssociation.image_id.is_(None),
            )
            .one_or_none()
        )

        if existing_association:
            logger.info(
                "No-image part-position association already exists | part_id=%s | position_id=%s",
                part.id,
                position_id,
            )
            return existing_association

        part_position_image_association = PartsPositionImageAssociation(
            part_id=part.id,
            position_id=position_id,
            image_id=None,
        )

        session.add(part_position_image_association)
        session.flush()

        logger.info(
            "Created part-position association without image | association_id=%s | part_id=%s | position_id=%s",
            getattr(part_position_image_association, "id", None),
            part.id,
            position_id,
        )
        return part_position_image_association

    def process_part_position_image(
        self,
        session: Session,
        part_number: str,
        position_id: int,
        image_title: str,
        base_image_path: str,
    ) -> Optional[int]:
        """
        Locate image, copy it to the database image folder, add it through the provided session,
        then create the part-position-image association.

        Returns:
            image_id if image was created/found, else None
        """
        logger.info(
            "Processing part-position-image | part_number=%s | position_id=%s | image_title=%s | base_image_path=%s",
            part_number,
            position_id,
            image_title,
            base_image_path,
        )

        image_file_path = self.find_image_in_subfolders(image_title, base_image_path)

        if not image_file_path:
            logger.error(
                "Image not found; aborting part-position-image processing | image_title=%s | base_image_path=%s",
                image_title,
                base_image_path,
            )
            return None

        os.makedirs(DATABASE_PATH_IMAGES_FOLDER, exist_ok=True)

        filename = os.path.basename(image_file_path)
        saved_image_path = os.path.join(DATABASE_PATH_IMAGES_FOLDER, filename)

        if not os.path.exists(saved_image_path):
            shutil.copy(image_file_path, saved_image_path)
            logger.info(
                "Copied image into database image folder | source=%s | destination=%s",
                image_file_path,
                saved_image_path,
            )
        else:
            logger.info(
                "Image already exists in database image folder | destination=%s",
                saved_image_path,
            )

        file_name_without_ext = os.path.splitext(os.path.basename(saved_image_path))[0]
        logger.info("Using image title derived from filename | image_title=%s", file_name_without_ext)

        image_id = self.get_or_create_image(
            session=session,
            title=file_name_without_ext,
            file_path=saved_image_path,
            position_id=position_id,
        )

        if image_id:
            logger.info(
                "Image available for association | image_title=%s | image_id=%s | position_id=%s",
                file_name_without_ext,
                image_id,
                position_id,
            )
            self.create_part_position_image_association(
                session=session,
                image_title=file_name_without_ext,
                position_id=position_id,
                image_id=image_id,
            )
        else:
            logger.error(
                "Failed to create or locate image | image_title=%s | saved_image_path=%s",
                file_name_without_ext,
                saved_image_path,
            )
            return None

        logger.info(
            "Finished part-position-image processing | part_number=%s | position_id=%s | image_title=%s",
            part_number,
            position_id,
            file_name_without_ext,
        )
        return image_id

    def get_or_create_image(
        self,
        session: Session,
        title: str,
        file_path: str,
        position_id: int,
    ) -> Optional[int]:
        """
        Create or reuse an Image row using the current orchestrator-owned session.

        This replaces the legacy helper that inserted images using its own DB pathway.

        Notes:
        - Uses title + file_path as the natural reuse key
        - Preserves ability to get image ID before orchestrator commit via session.flush()
        """
        logger.info(
            "Creating or finding image | title=%s | file_path=%s | position_id=%s",
            title,
            file_path,
            position_id,
        )

        existing_image = (
            session.query(Image)
            .filter(
                Image.title == title,
                Image.file_path == file_path,
            )
            .one_or_none()
        )

        if existing_image:
            logger.info(
                "Existing image found | image_id=%s | title=%s | file_path=%s",
                existing_image.id,
                title,
                file_path,
            )
            return existing_image.id

        description_value = title

        image = Image(
            title=title,
            description=description_value,
            file_path=file_path,
        )

        session.add(image)
        session.flush()

        logger.info(
            "Created image row | image_id=%s | title=%s | file_path=%s",
            image.id,
            title,
            file_path,
        )
        return image.id

    def create_part_position_image_association(
        self,
        session: Session,
        image_title: str,
        position_id: int,
        image_id: int,
    ) -> Optional[PartsPositionImageAssociation]:
        """
        Create the part_position_image association from image title and position.

        Returns the created association or existing association if already present.
        """
        try:
            logger.info(
                "Creating part-position-image association | image_title=%s | position_id=%s | image_id=%s",
                image_title,
                position_id,
                image_id,
            )

            part_number = image_title[:7]
            logger.info(
                "Derived part number from image title | image_title=%s | part_number=%s",
                image_title,
                part_number,
            )

            part = session.query(Part).filter(Part.part_number == part_number).one()
        except NoResultFound:
            logger.error(
                "No part found for image association | image_title=%s | derived_part_number=%s",
                image_title,
                image_title[:7],
            )
            return None

        existing_association = (
            session.query(PartsPositionImageAssociation)
            .filter(
                PartsPositionImageAssociation.part_id == part.id,
                PartsPositionImageAssociation.position_id == position_id,
                PartsPositionImageAssociation.image_id == image_id,
            )
            .one_or_none()
        )

        if existing_association:
            logger.info(
                "Part-position-image association already exists | association_id=%s | part_id=%s | position_id=%s | image_id=%s",
                getattr(existing_association, "id", None),
                part.id,
                position_id,
                image_id,
            )
            return existing_association

        part_position_image_association = PartsPositionImageAssociation(
            part_id=part.id,
            position_id=position_id,
            image_id=image_id,
        )

        session.add(part_position_image_association)
        session.flush()

        logger.info(
            "Created part-position-image association | association_id=%s | part_id=%s | position_id=%s | image_id=%s",
            getattr(part_position_image_association, "id", None),
            part.id,
            position_id,
            image_id,
        )
        return part_position_image_association

    # ==========================================================
    # INTERNAL HELPERS
    # ==========================================================
    def _get_bom_sheet_name(self, workbook) -> str:
        """
        Find the generated bom_* sheet in the target workbook.
        """
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("bom_"):
                return sheet_name
        raise KeyError("No sheet starting with 'bom_' was found in the workbook.")

    def _safe_str(self, value) -> str:
        """
        Convert worksheet values safely to string while treating None as empty string.
        """
        return "" if value is None else str(value)